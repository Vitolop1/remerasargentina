from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = ROOT / "REMERAS_SOCIOS.xlsx"
OUTPUT_PATH = ROOT / "src" / "data" / "catalog.json"


def round_if_number(value: Any, digits: int = 4) -> Any:
    if isinstance(value, (int, float)):
        return round(float(value), digits)
    return value


def export_catalog() -> None:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    settings_sheet = workbook.worksheets[0]
    order_sheet = workbook.worksheets[2]

    settings = {
        "exchangeRateArsPerUsd": round_if_number(settings_sheet["C8"].value, 2),
        "unitsPurchased": int(settings_sheet["C13"].value),
        "unitsForSale": int(settings_sheet["C14"].value),
        "reservedUnits": int(settings_sheet["C15"].value),
        "defaultSalePriceUsd": round_if_number(settings_sheet["C19"].value, 2),
        "defaultSalePriceArs": round_if_number(settings_sheet["C20"].value, 2),
    }

    order_lines = []
    for row in order_sheet.iter_rows(min_row=3, values_only=True):
        line_number = row[0]
        if line_number == "TOTAL":
            break
        if not isinstance(line_number, int):
            continue

        order_lines.append(
            {
                "lineNumber": line_number,
                "name": row[1],
                "size": row[2],
                "quantity": int(row[4]),
            }
        )

    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceWorkbook": WORKBOOK_PATH.name,
        "settings": settings,
        "orderLines": order_lines,
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Exported {len(order_lines)} public order lines to {OUTPUT_PATH}")


if __name__ == "__main__":
    export_catalog()
